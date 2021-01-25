#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time    : 2021/1/25 13:04
# @Author  : yuwenli


from openpyxl import Workbook, load_workbook
import random

wb = Workbook()
ws = wb.active
dest_filename = "test.xlsx"
sheet_name = "test1"
ws.title = sheet_name
ws['A1'] = "TEST01"
wb.save(dest_filename)

wb = load_workbook(dest_filename)
ws = wb[sheet_name]
# print(ws['A1'])

wb.create_sheet("test02")

ws2 = wb["test02"]
for i in range(1, 3):
    for j in range(1, 5):
        ws2.cell(row=i, column=j, value=random.randint(1, 50))
# 按行读取
for row in ws2.rows:
    for cell in row:
        print(cell.value)
# 按列读取
for column in ws2.columns:
    for cell in column:
        print(cell.value)

print("最大的行：%d " % ws2.max_row)
print("最大的列：%d " % ws2.max_column)
# 第2行，输出为一个tuple
# print(ws2[2])
# print(tuple(ws2.rows))
print(ws2['A1':'B2'])
# 输出为tuple
print(type(ws2['A1':'B2']))

for i in range(1, 10):
    for j in range(1, 5):
        ws2.cell(row=i, column=j).value = random.randint(1, 50)

wb.save(dest_filename)
