#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time    : 2021/1/24 20:38
# @Author  : yuwenli

"""
注意：当一个工作表被创建是，其中不包含单元格cell。只有当单元格被获取是才被创建。这种方式我们不会创建我们从不会使用的单元格，从而减少了内存消耗。
警告：由于上述特性，你如果遍历了单元格而非想要使用它们也将会在内存当中创建。
w = workbook()创建workbook，实例化
w.active 获取默认创建的工作簿
wb = load_workbook('ecxelname.xlsx') 打开已有的workbook
w.title="" 工作簿命名
# 创建表
w.create_sheet(title="") 创建工作簿，默认插入最后一个位置
w.create_sheet("mysheet")
ws = w.create_sheet("mysheet",0) 创建工作簿，默认插入第一个位置
# 选择表
1.根据title名称选择表（sheet名称可以作为key进行索引）
w1 = w['title1'] 根据title名称获取工作簿
w1 = w.get_sheet_by_name("mysheet")
# 查看表名
1.显示所有表名
workbook.sheetnames 获取所有worksheet
2.遍历所有表
for sheet in wb:
    print(sheet.title)

# 存储数据
单元格赋值，直接分配到某个单元格
ws['F1']="" 根据索引赋值
F1 = ws['F1'] 获取某个单元格值
附加行：
ws.append([1,2,3])从第一列开始附加，从最左边开始
类型自动转换：
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

# 访问单元格
1. 单个单元格访问
A1 = ws['A1']
2.row 行，column列
d = ws.cell(row=1,column=1,value=10)

2. 创建空的cell，不赋值
for i in range(1,101):
  for j in range(1,101):
    ws.cell(row=i,column=j)

3. 访问生产100X100单元格并赋值（可不赋值）
for i in range(1,101):
  for j in range(1,101):
    ws2.cell(row=i,column=j,value=i*j)

4.使用切片访问单元格
cell_range=['A2','D6']
for cel in cell_range: #元组数据
 for cl in cel:  #遍历元组数据
   cl.value=random.randint(1,50)  #随机赋值

5.使用列访问单元格
colC= ws['C'] #c列
for cel in colC:
  cel.value=random.randin(1,50)

col_range=ws['C:D']
for cel in col_range:
  for cl in cel:
   cl.value=random.randint(1,50)

6.使用行访问单元格（行访问，直接行数）
#第10行
row10=ws[10]
第2到8行
row_range=ws[2:9]

7.指定行列访问
#使用Workbook.iter_rows()方法访问单元格
for row in ws.iter_rows(min_row=1,max_col=4,max_row=6):
    for cel in row:
        cel.value=random.randint(1,30)

# 指定行-行
for row in  ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in  row:
        print(cell)

# 指定列-列
#使用Workbook.iter_cols()方法访问单元格(在只读模式下不可用)
for row in ws.iter_cols(min_row=1,max_col=4,max_row=6):
    for cel in row:
        cel.value=random.randint(1,30)
for row in  ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in  row:
        print(cell)

8. 遍历所有
遍历所有行或者列可以使用Worksheet.rows /Worksheet.columns
ws['C5']='hello'
print(tuple(ws.rows) #访问A1~C5范围的单元格
tuple(ws.columns)
tuple(ws.rows)

# 保存数据
wb = Workbook()
wb.save('balances.xlsx')

# 改变sheet的color
ws.sheet_properties.tabColor = "1072BA"

# 获取最大行，最大列
sheet.max_row
sheet.max_column

# 获取每一行，每一列
sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
sheet.columns类似，不过里面是每个tuple是每一列的单元格。
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)

#根据列的数字返回字母
get_column_letter(1):第一列的字母为A
# 根据字母返回列的数字
column_index_from_string('A') #1

# 删除工作表
wb.remove(sheet)
del wb[sheet]

"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 创建workbook
wb = Workbook()
dest_filename = 'empty_book.xlsx'
# 创建workbook的时候会默认创建出一个worksheet
# 使用active就可以得到这个默认创建的worksheet
ws1 = wb.active
# 给指定的worksheet赋予标题
ws1.title = "range names"
# 39行
for row in range(1, 40):
    ws1.append(range(600))
# 创建新的工作簿，并指定工作簿名称title
ws2 = wb.create_sheet(title="Pi")
# 给某个单元格赋值(根据索引赋值）
ws2['F5'] = 3.14
# 创建新的工作簿, 默认插在工作簿末尾
ws3 = wb.create_sheet(title="Data")
ws4 = wb.create_sheet(0)  # 插入在工作簿的第一个位置
# 给工作簿命名
ws4.title = "test4"
# row 行：10行
for row in range(10, 20):
    # col列：28列
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
